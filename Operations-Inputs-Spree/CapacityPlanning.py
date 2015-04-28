# -*- coding: utf-8 -*-
"""
Created on Fri Jan 09 14:13:28 2015
"""
import pandas as pd
from pandas import DataFrame
import gspread
import numpy as np
#import time
from datetime import date
from pandas import ExcelWriter
from openpyxl.reader.excel import load_workbook

#Email Import Library
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import Encoders

today=date.today()

pswd = raw_input("Please enter your password:")
email_address= raw_input("Please enter your email address:")

#Run Totalxlxs Script for the latest stock count info
import sys
sys.path.append('Y:\\SUPPLY CHAIN\\Python Scripts\\00_SharedFunctions')
import Totalxlsx
DocName = 'Rolling Stock'
HistoryPath = 'Y:\\SUPPLY CHAIN\\Stock Count\\All handovers'
SavePath = 'Y:\\SUPPLY CHAIN\\Python Scripts\\02_StockCount'
Totalxlsx.data_total( DocName, HistoryPath, SavePath )

#Create rolling stock dataframe
R_Stock=pd.read_excel('Y:\\SUPPLY CHAIN\\Python Scripts\\02_StockCount\\Rolling Stock.xlsx',index_col=None, na_values=['NA'])

#Create Category_Conversion Table for Reporting
Product_Table=pd.read_csv('C:\Users\Laurie.Richardson\Operations-Inputs-Spree\Operations-Inputs-Spree\Product Conversion Table 9 Feb.csv', na_values=['NA'])
Category_Conversion=Product_Table.set_index('DETAILED CATEGORY')['CATEGORY'].to_dict()

#Import Post-It Workflow data
c = gspread.Client(auth=(email_address,pswd))
c.login()
c.open('Stock Processing')
sht=c.open('Stock Processing')
worksheet = sht.worksheet('Post It_Workflow')
info=worksheet.get_all_values()
headers=info.pop(0)
df1=DataFrame(data=info,columns=headers)

#Import Epping Receiving Data
d=gspread.Client(auth=(email_address,pswd))
d.login()
d.open('Epping Receiving Report')
sht=d.open('Epping Receiving Report')
worksheet = sht.worksheet('Booked')
info=worksheet.get_all_values()
headers=info.pop(0)
df2=DataFrame(data=info,columns=headers)

#Creating the last rolling 14 day PO received count dataframe
DailyReceived=df2[['Date received','POs']] 
DailyReceived['Date received']=pd.to_datetime(DailyReceived['Date received'],coerce=True)
x=DailyReceived.set_index(pd.DatetimeIndex(DailyReceived['Date received']))
y=x['POs']
grouped=y.groupby(level=0)
x=grouped.count()
Rolling_14_Day_Receives=x.ix[-14:]


#Creating the number of PO's processed per day 
DailyQC=df1[['POs','Supplier_Name','Sort and Count Start Date:','QC Start Date:','QC End Date:','Sort and Count Team','QCTeamName']] 
DailyQC['Sort and Count Start Date:']=pd.to_datetime(DailyQC['Sort and Count Start Date:'],coerce=True)
DailyQC['QC Start Date:']=pd.to_datetime(DailyQC['QC Start Date:'],coerce=True)
DailyQC['QC End Date:']=pd.to_datetime(DailyQC['QC End Date:'],coerce=True)
DailyQC= DailyQC[pd.notnull(DailyQC['QC Start Date:'])]
DailyQC= DailyQC[pd.notnull(DailyQC['Sort and Count Start Date:'])]
DailyQC['Days to process a PO']=DailyQC['QC End Date:']-DailyQC['QC Start Date:']

a=DailyQC.set_index(pd.DatetimeIndex(DailyQC['QC Start Date:']))
a['Days to process a PO']=a['Days to process a PO'].apply(lambda x: x / np.timedelta64(1,'D'))
a=a.fillna(-1)
a['Days to process a PO'].astype(int)
b=a['Days to process a PO']

groupb=b.groupby(level=0)
groupedQC=groupb.count()
Rolling_14_Day_QC=groupedQC.ix[-14:]

#Compare the PO's received per day to the number of POs QCd per day
z=pd.concat([Rolling_14_Day_QC,Rolling_14_Day_Receives], join='outer',axis=1)
z = z.rename(columns = {'Days to process a PO':'POs processed by QC per day','POs':'POs Received per day from suppliers'})

#How long a PO stays within the system from arrival to end of QC-Rolling 14 Week Period
TotalTime=pd.merge(DailyReceived,DailyQC,how='left',on='POs')
TotalTime=TotalTime[pd.notnull(TotalTime['Date received'])]
TotalTime=TotalTime.drop_duplicates(['POs'],take_last='True')
TotalTime['PO_Days_from_Receiving_to_QC_Complete']=TotalTime['QC End Date:']-TotalTime['Date received']
TotalTime['PO_Days_from_Receiving_to_QC_Complete']=TotalTime['PO_Days_from_Receiving_to_QC_Complete'].apply(lambda x: x / np.timedelta64(1,'D'))
TotalTime['PO_Days_from_Receiving_to_QC_Complete']=TotalTime['PO_Days_from_Receiving_to_QC_Complete'].fillna(0)
TotalTime['PO_Days_from_Receiving_to_QC_Complete'].astype(int)
System_Days=TotalTime[['Date received','Supplier_Name','PO_Days_from_Receiving_to_QC_Complete']]

System_Days_Filter1=System_Days[System_Days.Supplier_Name !='Samples']
System_Days_Filter_2=System_Days_Filter1[System_Days_Filter1.Supplier_Name !='samples']

System_Days_Drop=System_Days_Filter_2[['Date received','PO_Days_from_Receiving_to_QC_Complete']]
System_Total=System_Days_Drop.query('PO_Days_from_Receiving_to_QC_Complete>0')
Summ=System_Total.set_index(pd.DatetimeIndex(System_Total['Date received']))
Summ_1=Summ[['PO_Days_from_Receiving_to_QC_Complete']]
PO_System_Time=Summ_1.resample('W',how='mean')
PO_System_Time.index.names = ['Week of PO Arrival']
PO_System_Time_Final=PO_System_Time.rename(columns = {'PO_Days_from_Receiving_to_QC_Complete':'Ave. Number of days taken to process a PO for the week'})
PO_System_Time_Final_1=PO_System_Time_Final.ix[-14:]

# Number of Units processed per rolling 14 # period
R_Stock_DF=R_Stock[['Category','Date', 'PO','Qty Counted','Team']]
R_Stock_DF['Category']=R_Stock_DF['Category'].fillna(value='Unknown')
R_Stock_DF['High Level Category']=R_Stock_DF['Category'].map(Category_Conversion)
R_Stock_DF['Date']=pd.to_datetime(R_Stock_DF['Date'],coerce=True)
R_Stock_Count=R_Stock_DF.set_index(pd.DatetimeIndex(R_Stock_DF['Date']))

# By Rolling 14 Days
R_Stock_Count_Day=R_Stock_Count.resample('B',how='sum')
R_Stock_Count_Day.index.names = ['Date Counted']
R_Stock_Count_Day.ix[-14:]

# By Rolling 14 Weeks
R_Stock_Count_Week=R_Stock_Count.resample('W',how='sum')
R_Stock_Count_Week.index.names = ['Week Counted']
R_Stock_Count_Week.ix[-14:]

#Number of units processed by QC Team-Rolling 14 day period
R_Stock_Count_Group_D=R_Stock_DF.set_index('Date').groupby('Team').resample('B', how='sum')
R_Stock_Count_Group_D=R_Stock_Count_Group_D.reset_index()
R_Stock_Count_Group_D_Pivot=pd.pivot_table(R_Stock_Count_Group_D, values='Qty Counted', index=['Date'],columns=['Team'],aggfunc=np.sum).ix[-14:]

#Number of units processed by QC Team-Rolling 14 week period
R_Stock_Count_Group_W=R_Stock_DF.set_index('Date').groupby('Team').resample('W', how='sum')
R_Stock_Count_Group_W=R_Stock_Count_Group_W.reset_index()
R_Stock_Count_Group_W_Pivot=pd.pivot_table(R_Stock_Count_Group_W, values='Qty Counted', index=['Date'],columns=['Team'],aggfunc=np.sum).ix[-14:]

#Number of units processed per Category-Rolling daily period
R_Stock_Count_Group_W_Cat=R_Stock_DF.set_index('Date').groupby(['Category','High Level Category','Team']).resample('B', how='sum')
R_Stock_Count_Group_W_Cat=R_Stock_Count_Group_W_Cat.reset_index()
R_Stock_Count_Group_W_Cat_Pivot=pd.pivot_table(R_Stock_Count_Group_W_Cat, values='Qty Counted', index=['Date','High Level Category','Category'],columns=['Team'],aggfunc=np.sum)
R_Stock_Count_Group_W_Cat_Pivot=R_Stock_Count_Group_W_Cat_Pivot.fillna(value=0)
R_Stock_Count_Group_W_Cat_Pivot_Filter=R_Stock_Count_Group_W_Cat_Pivot.sort(ascending=False)
R_Stock_Count_Group_W_Cat_Pivot_Filter2=R_Stock_Count_Group_W_Cat_Pivot_Filter.query('A>0 or B>0')

#Write to Excel Spreadsheet
#writer3 = ExcelWriter('C:\Users\Laurie.Richardson\Operations-Inputs-Spree\Operations-Inputs-Spree\SpreeQCTeamOutput\SpreeQCTeamOutput ' + str(today) + '.xlsx')
writer3 = ExcelWriter('Spree QCTeam Output ' + str(today) + '.xlsx')

#PO Detailed Information in Dataframe to Excel
z.to_excel(writer3, 'PO Detail', startrow = 3, startcol=0,na_rep=0)
PO_System_Time_Final_1.to_excel(writer3, 'PO Detail', startrow = 3, startcol=4,na_rep=0, float_format='%.2f')

#Unit level data to Excel
R_Stock_Count_Day.ix[-14:].to_excel(writer3, 'Units Counted', startrow = 3, startcol=0,na_rep=0)
R_Stock_Count_Week.ix[-14:].to_excel(writer3,'Units Counted', startrow = 3, startcol=3,na_rep=0)

#QC team data to Excel
R_Stock_Count_Group_D_Pivot.to_excel(writer3, 'QC_Team Output', startrow = 3, startcol=0,na_rep=0)
R_Stock_Count_Group_W_Pivot.to_excel(writer3, 'QC_Team Output', startrow = 3, startcol=6,na_rep=0)

#Category Data to Excel
R_Stock_Count_Group_W_Cat_Pivot_Filter2.to_excel(writer3, 'Category Output', startrow = 2, startcol=0,na_rep=0)

#Writer
workbook =  writer3.book
workbook1 = writer3.book
workbook2 = writer3.book
workbook3=  writer3.book

#Formatting for the PO and Units Detail and QC Team Output
title = workbook.add_format({'bold':True, 'size':14})
header = workbook.add_format({'size':12, 'underline':True, 'font_color':'green'})

#Formatting and text for the PO Detail
worksheet = writer3.sheets['PO Detail']
worksheet.write('A1','Purchase Order Details ' + str(today), title)
worksheet.write('A3','Purchase Orders Processed and Received', header)
worksheet.write('E3','Average Number of Days to Process a Purchase Order',header)

worksheet.set_column('A:A', 18)
worksheet.set_column('B:C', 34)
worksheet.set_column('D:D', 12)
worksheet.set_column('E:E', 18)
worksheet.set_column('F:F', 50)

#Formatting and text for the Units Count 
worksheet1 = writer3.sheets['Units Counted']
worksheet1.write('A1','Units Counted Details ' + str(today), title)
worksheet1.write('A3','Daily Count', header)
worksheet1.write('D3','Weekly Count',header)

worksheet1.set_column('A:A', 18)
worksheet1.set_column('B:B', 18)
worksheet1.set_column('C:C', 12)
worksheet1.set_column('D:D', 18)
worksheet1.set_column('E:E', 18)

#Formatting and text for the QC Team Output 
worksheet2 = writer3.sheets['QC_Team Output']
worksheet2.write('A1','QC Team Output ' + str(today), title)
worksheet2.write('A3','Daily Team Output', header)
worksheet2.write('G3','Weekly Team Output',header)

worksheet2.set_column('A:A', 18)
worksheet2.set_column('B:E', 15)
worksheet2.set_column('F:F', 12)
worksheet2.set_column('G:G', 18)
worksheet2.set_column('H:K', 15)

#Formatting and text for the Category Output 
worksheet3 = writer3.sheets['Category Output']
worksheet3.write('A1','Daily Category Output by QC Team ' + str(today), title)

worksheet3.set_column('A:A', 14)
worksheet3.set_column('B:B', 50)
worksheet3.set_column('C:C', 28)
worksheet3.set_column('D:F', 10)

writer3.save()

#format Excel Spreadsheet (use when data already written to excel)
#wb = load_workbook('C:\Users\Laurie.Richardson\Operations-Inputs-Spree\Operations-Inputs-Spree\SpreeQCTeamOutput\SpreeQCTeamOutput ' + str(today) + '.xlsx')
wb = load_workbook('Spree QCTeam Output ' + str(today) + '.xlsx')

#For PO Detail
ws = wb.worksheets[0]

cellsA = [ws['B4'],ws['C4'],ws['F4']]
for cell in cellsA:
    cell.style.alignment.wrap_text = True 
    
cellsB = ws['A4':'I22']
for row in cellsB:
    for cell in row:
        cell.style.alignment.horizontal = 'center'
        
cellsC = ws['A5':'A22']
for row in cellsC:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

cellsD = ws['E6':'E19']
for row in cellsD:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

#For Unit Level Detail
ws1 = wb.worksheets[1]

cellsD = [ws1['B4'],ws1['E4']]
for cell in cellsD:
    cell.style.alignment.wrap_text = True 
    
cellsE = ws1['A4':'E22']
for row in cellsE:
    for cell in row:
        cell.style.alignment.horizontal = 'center'
        
cellsF = ws1['A6':'A22']
for row in cellsF:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

cellsG = ws1['D6':'D19']
for row in cellsG:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

#For QC Team Output
ws2 = wb.worksheets[2]

cellsH = [ws2['A4'],ws2['G4']]
for cell in cellsH:
    cell.style.alignment.wrap_text = True 
    
cellsI = ws2['A4':'K25']
for row in cellsI:
    for cell in row:
        cell.style.alignment.horizontal = 'center'
        
cellsJ = ws2['A6':'A22']
for row in cellsJ:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

cellsK = ws2['G6':'G19']
for row in cellsK:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

#For Category Output
ws3 = wb.worksheets[3]
    
cellsM = ws3['A4':'G180']
for row in cellsM:
    for cell in row:
        cell.style.alignment.horizontal = 'center'
        
cellsN = ws3['A5':'A180']
for row in cellsN:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'


wb.save('Spree QCTeam Output ' + str(today) + '.xlsx')

#wb.save('C:\Users\Laurie.Richardson\Operations-Inputs-Spree\Operations-Inputs-Spree\SpreeQCTeamOutput\SpreeQCTeamOutput ' + str(today) + '.xlsx')


#Details to send Email to 
msg = MIMEMultipart()
today=date.today()
doc_name= 'Spree QCTeam Output '  
message = 'Data on QC Team Performance'        
part ='Spree QCTeam Output ' + str(today) + '.xlsx'

#'C:\Users\Laurie.Richardson\Operations-Inputs-Spree\Operations-Inputs-Spree\SpreeQCTeamOutput\SpreeQCTeamOutput

today = date.today()
urlFile = open("MailList.txt", "r+")
MailList = [i.strip() for i in urlFile.readlines()]    
        
fromEmail = email_address 

#Create message
msg['Subject'] = str(doc_name) + str(today)
msg['From'] = fromEmail
body = message
content = MIMEText(body, 'plain')
msg.attach(content)
        
#Create attachment        
filename = str(part)
f = file(filename)
attachment = MIMEText(f.read())
attachment.set_payload(open(part, 'rb').read())
Encoders.encode_base64(attachment)
attachment.add_header('Content-Disposition', 'attachment', filename=filename)           
msg.attach(attachment)
        
#Call server and send email      
mailServer = smtplib.SMTP('smtp.gmail.com', 587)
mailServer.set_debuglevel(1)
mailServer.ehlo()
mailServer.starttls()
mailServer.ehlo()
mailServer.login(email_address, pswd)
mailServer.ehlo()
mailServer.sendmail(fromEmail, MailList, msg.as_string())
mailServer.quit()
        
