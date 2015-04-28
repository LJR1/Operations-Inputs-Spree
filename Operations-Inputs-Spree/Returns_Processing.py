# -*- coding: utf-8 -*-
"""
Created on Wed Mar 11 11:12:48 2015

"""
import pandas as pd
from pandas import DataFrame
import gspread
import numpy as np
#import time
from datetime import date
from pandas import ExcelWriter
from openpyxl.reader.excel import load_workbook

#Email Import Library
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import Encoders

today=date.today()

pswd = raw_input("Please enter your google password:")
email_address= raw_input("Please enter your email address:")

#Import QC-Returns data
c = gspread.Client(auth=(email_address,pswd))
c.login()
c.open('QC returns')
sht=c.open('QC returns')
worksheet = sht.worksheet('Spree receiving returns - Master')

info=worksheet.get_all_values()
headers=info.pop(0)
df=DataFrame(data=info,columns=headers)

d = gspread.Client(auth=(email_address,pswd))
d.login()
d.open('QC returns')
sht=d.open('QC returns')
worksheet = sht.worksheet('Returns - Master 2015')
info=worksheet.get_all_values()
headers=info.pop(0)
df1=DataFrame(data=info,columns=headers)

#Number of returns processed per day per QC Agent (Rolling 14 day)

ReturnsQC=df1[['Date Spree QC processed the return','QC agent name']] 
ReturnsQC['Date Spree QC processed the return']=pd.to_datetime(ReturnsQC['Date Spree QC processed the return'],coerce=True)
ReturnsQC['Returns QCd per agent']=1

ReturnsQC['QC agent name'] = ReturnsQC['QC agent name'].str.title().str.strip('-').dropna(how='all')

ReturnsQCPivot=pd.pivot_table(ReturnsQC, values='Returns QCd per agent',index=['Date Spree QC processed the return'],columns=['QC agent name'],aggfunc=np.sum).ix[-14:]
#ReturnsQCPivot=ReturnsQCPivot.drop(['Caroline','Genevieve','N/A'],axis=1)

#Number of Returns processed per day by Order Number per QC Agent (Rolling 14 day)
ReturnsQC_Order=df1[['Date Spree QC processed the return','QC agent name','Order Number']]
ReturnsQC_Order['Date Spree QC processed the return']=pd.to_datetime(ReturnsQC_Order['Date Spree QC processed the return'],coerce=True)

ReturnsQC_Order['QC agent name'] = ReturnsQC_Order['QC agent name'].str.title().str.strip('-')

ReturnsQC_Order_Group=ReturnsQC_Order.groupby(['Date Spree QC processed the return','QC agent name'])['Order Number'].value_counts()
ReturnsQC_Order_Group_Counts=pd.DataFrame(ReturnsQC_Order_Group).reset_index()
ReturnsQC_Order_Group_Counts['Returns Processed by Order Level']=1
ReturnsQC_Order_Group_Counts_Final=ReturnsQC_Order_Group_Counts[['Date Spree QC processed the return','QC agent name','Returns Processed by Order Level']]
ReturnsQC_Order_Group_Counts_Pivot=pd.pivot_table(ReturnsQC_Order_Group_Counts_Final,values='Returns Processed by Order Level',index='Date Spree QC processed the return',columns=['QC agent name'],aggfunc=np.sum)
ReturnsQC_Order_Group_Counts_Pivot=ReturnsQC_Order_Group_Counts_Pivot.ix[-14:]
#ReturnsQC_Order_Group_Counts_Pivot=ReturnsQC_Order_Group_Counts_Pivot.drop(['Caroline','Genevieve','N/A'],axis=1).ix[-14:]

#Number of Returns received and processed on a daily basis (Rolling 14 day)
OrdersReceived=df[['Date ','Couriers waybill number']]
OrdersReceived.rename(columns={'Date ':'Arrival date of Return','Couriers waybill number':'Waybill Number'},inplace=True)
OrdersReceived['Arrival date of Return']=pd.to_datetime(OrdersReceived['Arrival date of Return'],coerce=True)
OrdersReceived['Returns that arrived']=1
OrdersReceivedTime=OrdersReceived.set_index(pd.DatetimeIndex(OrdersReceived['Arrival date of Return']))
OrdersReceivedIndex=OrdersReceivedTime['Returns that arrived']
grouped=OrdersReceivedIndex.groupby(level=0)
OrdersReceivedFinal=grouped.count()
Rolling_14_Day_Receives=OrdersReceivedFinal.ix[-14:]

ReturnsQCIndex=ReturnsQC_Order_Group_Counts_Final[['Date Spree QC processed the return','Returns Processed by Order Level']]
ReturnsQCIndex['Date Spree QC processed the return']=pd.to_datetime(ReturnsQCIndex['Date Spree QC processed the return'],coerce=True)
ReturnsQCIndexed=ReturnsQCIndex.set_index(pd.DatetimeIndex(ReturnsQCIndex['Date Spree QC processed the return']))
ReturnsQCIndexed1=ReturnsQCIndexed['Returns Processed by Order Level']
grouped1=ReturnsQCIndexed1.groupby(level=0)
ReturnsQCFinal=grouped1.count()
Rolling_14_Day_QC=ReturnsQCFinal.ix[-14:]

Output=pd.concat([Rolling_14_Day_Receives,Rolling_14_Day_QC], join='outer',axis=1)

#Write to Excel Spreadsheet

writer3 = ExcelWriter('Returns Processing ' + str(today) + '.xlsx')

#Number of returns processed per day per QC Agent (Rolling 14 day) to Excel

ReturnsQCPivot.to_excel(writer3, 'Returns Team Output', startrow = 3, startcol=0,na_rep=0)
ReturnsQC_Order_Group_Counts_Pivot.to_excel(writer3, 'Returns Team Output', startrow = 21, startcol=0,na_rep=0)
Output.to_excel(writer3,'Arrivals vs QCd',startrow = 3, startcol=0,na_rep=0)


#Writer
workbook =  writer3.book
workbook1 = writer3.book

#Formatting for the Returns Processed per day per QC Agent
title = workbook.add_format({'bold':True, 'size':14})
header = workbook.add_format({'size':12, 'underline':True, 'font_color':'green'})

#Formatting and text for the Returns Team Processing Output
worksheet = writer3.sheets['Returns Team Output']
worksheet.write('A1','Returns Output per QC Agent ' + str(today), title)
worksheet.write('A3','Returns Output by Units',header)
worksheet.write('A21','Returns Output by Orders',header)

worksheet.set_column('A:A', 34)
worksheet.set_column('B:I', 14)

#Formatting and text for the Returns Processing Output vs Arrivals
worksheet1 = writer3.sheets['Arrivals vs QCd']
worksheet1.write('A1','Return Arrivals vs QCd ' + str(today), title)
worksheet1.write('A3','Returns Arrivals vs QCd by Order Level',header)

worksheet1.set_column('A:D',45)

writer3.save()

#format Excel Spreadsheet (use when data already written to excel)
wb = load_workbook('Returns Processing ' + str(today) + '.xlsx')

#For Returns Team Output
ws = wb.worksheets[0]

#cellsA = [ws['A4':'I4']]
#for cell in cellsA:
 #   cell.style.alignment.wrap_text = True 
    
cellsB = ws['A3':'I50']
for row in cellsB:
    for cell in row:
        cell.style.alignment.horizontal = 'center'
        
cellsC = ws['A5':'A50']
for row in cellsC:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

#For Returns Arrival vs QCd
ws1 = wb.worksheets[1]

cellsD = ws1['A3':'E30']
for row in cellsD:
    for cell in row:
        cell.style.alignment.horizontal = 'center'
        
cellsE = ws1['A3':'A30']
for row in cellsE:
    for cell in row:
        cell.style.number_format.format_code = 'dd/mm/yyyy'

wb.save('Returns Processing ' + str(today) + '.xlsx')

#Details to send Email to 

msg = MIMEMultipart()

today=date.today()
        
doc_name= 'Returns Processing '  
message = 'Data on Returns Performance'        
part = 'Returns Processing ' + str(today) + '.xlsx'
today = date.today()
urlFile = open("MailListReturns.txt", "r+")
MailList = [i.strip() for i in urlFile.readlines()]    
        
fromEmail = email_address

#create message
msg['Subject'] = str(doc_name) + str(today)
msg['From'] = fromEmail
#msg['To'] = ', '.join(MailList)
body = message
content = MIMEText(body, 'plain')
msg.attach(content)
        
#create attachment        
filename = str(part)
f = file(filename)
attachment = MIMEText(f.read())
attachment.set_payload(open(part, 'rb').read())
Encoders.encode_base64(attachment)
attachment.add_header('Content-Disposition', 'attachment', filename=filename)           
msg.attach(attachment)
        
#call server and send email      
mailServer = smtplib.SMTP('smtp.gmail.com', 587)
mailServer.set_debuglevel(1)
mailServer.ehlo()
mailServer.starttls()
mailServer.ehlo()
mailServer.login(email_address, pswd)
mailServer.ehlo()
mailServer.sendmail(fromEmail, MailList, msg.as_string())
mailServer.quit()
